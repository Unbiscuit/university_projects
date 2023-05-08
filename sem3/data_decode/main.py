from openpyxl import load_workbook


def get_numbers():

    with open('numbers.dict', 'w') as file:
        for i in range(10):
            for j in range(10):
                for k in range(10):
                    for m in range(10):
                        for n in range(10):
                            for b in range(10):
                                for a in range(10):
                                    for l in range(10):
                                        for v in range(10):
                                            file.write(f'89{i}{j}{k}{m}{n}{b}{a}{l}{v}')
                                            file.write('\n')


def get_hashes():

    wb = load_workbook(filename='student_v.1.03.xlsx')
    ws = wb.active
    with open('hashes.txt', 'w') as file:
        for cell in range(2, ws.max_row + 1):
            file.write(ws[f'A{cell}'].value)
            file.write('\n')


def find_shift_mail(ws, index):

    alphabet = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z']
    server = list(ws[f'B{index}'].value)
    alphabet = alphabet + alphabet
    shift = 0

    while True:

        for j in range(len(server)):
            if server[j] in alphabet:
                server[j] = alphabet[alphabet.index(server[j]) + 1]

        shift += 1

        if server[-3] + server[-2] + server[-1] == 'net':
            return shift

        if server[-3] + server[-2] + server[-1] == 'biz':
            return shift

        if server[-3] + server[-2] + server[-1] == 'org':
            return shift

        if server[-3] + server[-2] + server[-1] == 'com':
            return shift

        if server[-4] + server[-3] + server[-2] + server[-1] == 'info':
            return shift


def find_address_shift(ws, index):

    alphabet = ['а', 'б', 'в', 'г', 'д', 'е', 'ж', 'з', 'и', 'й', 'к', 'л', 'м', 'н', 'о', 'п', 'р', 'с', 'т', 'у', 'ф', 'х', 'ц', 'ч', 'ш', 'щ', 'ъ', 'ы', 'ь', 'э', 'ю', 'я']
    upper_alphabet = ['А', 'Б', 'В', 'Г', 'Д', 'Е', 'Ж', 'З', 'И', 'Й', 'К', 'Л', 'М', 'Н', 'О', 'П', 'Р', 'С', 'Т', 'У', 'Ф', 'Х', 'Ц', 'Ч', 'Ш', 'Щ', 'Ъ', 'Ы', 'Ь', 'Э', 'Ю', 'Я']
    address = list(ws[f'C{index}'].value)
    alphabet = alphabet + alphabet
    upper_alphabet = upper_alphabet + upper_alphabet
    shift = 0

    while True:

        for j in range(len(address)):

            if address[j] in alphabet:
                address[j] = alphabet[alphabet.index(address[j]) + 1]
            if address[j] in upper_alphabet:
                address[j] = upper_alphabet[upper_alphabet.index(address[j]) + 1]

        shift += 1

        address = address[::-1]
        tail = ''

        for i, element in enumerate(address):
            if element == ' ' or element == '':
                break
            tail = tail + element

        address = address[::-1]
        tail = tail[::-1]

        if 'кв' in tail:
            return shift, ''.join(address)


def decode_mail(ws, index, shift):

    alphabet = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z']
    alphabet = alphabet + alphabet
    encrypted_email = list(ws[f'B{index}'].value)
    decrypted_email = []

    for element in encrypted_email:

        if element in alphabet:
            decrypted_email.append(alphabet[alphabet.index(element) + shift])
        else:
            decrypted_email.append(element)

    return ''.join(decrypted_email)


def main():
    # get_hashes()
    # get_numbers()
    wb = load_workbook(filename='student_v.1.03.xlsx')
    ws = wb.active
    ws['D1'] = 'сдвиг почты'
    ws['E1'] = 'сдвиг адреса'
    with open('cracked.txt', 'r') as file:
        text = file.readlines()
        for i in range(2, ws.max_row + 1):
            string = text[i - 2]
            string = string[string.index(':') + 1:].replace('\n', '')
            ws[f'A{i}'] = string
            shift = find_shift_mail(ws, i)
            ws[f'D{i}'] = shift
            ws[f'B{i}'] = decode_mail(ws, i, shift)
            shift, address = find_address_shift(ws, i)
            ws[f'C{i}'] = address
            ws[f'E{i}'] = shift

    wb.save(filename="changed_student_3.xlsx")


if __name__ == '__main__':
    main()

