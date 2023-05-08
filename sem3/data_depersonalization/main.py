from openpyxl import load_workbook
import random


def mask_email(email):
    email = list(email)
    index = 0
    while email[index] != '@':
        if index == 0:
            email[index] = 'X'
        else:
            email[index] = ''
        index += 1
    return ''.join(email)

"""
def mask_ip(ip):
    ip = list(ip)
    ip.append('.')
    section = 0
    index = 0
    for element in ip:
        if element == '.':
            section += 1
        if section >= 2 and section != 4 and element == '.':
            shift = index + 1
            while ip[shift] != '.':
                if ip[shift - 1] != 'X' and ip[shift - 1] != '':
                    ip[shift] = 'X'
                else:
                    ip[shift] = ''
                shift += 1
        index += 1
    ip[-1] = ''
    return ''.join(ip)
"""


def alias_of_site(platform, codes, platforms_for_masking):
    index = platforms_for_masking.index(platform)
    code = codes[index]
    return code


def local_for_amount_of_ads(amount_of_ads):
    if amount_of_ads < 10:
        return '1'
    if 10 <= amount_of_ads < 20:
        return '2'
    if 20 <= amount_of_ads < 30:
        return '3'
    if 30 <= amount_of_ads < 40:
        return '4'
    if 40 <= amount_of_ads < 50:
        return '5'
    if 50 <= amount_of_ads < 60:
        return '6'
    if 60 <= amount_of_ads < 70:
        return '7'
    if 70 <= amount_of_ads < 80:
        return '8'
    if 80 <= amount_of_ads < 90:
        return '9'
    if 90 <= amount_of_ads <= 100:
        return '10'


def local_for_adv_time(adv_time):
    adv_time = list(adv_time)
    index = 0
    mins = ''
    while adv_time[index] != ':':
        mins = mins + adv_time[index]
        index += 1
    if int(mins) <= 60:
        return 'недолго'
    else:
        return 'долго'


def local_for_product(product):
    product = list(product)
    index = 0
    thing = ''
    while product[index] != ' ':
        thing = thing + product[index]
        index += 1
    if thing == 'шуба' or thing == 'шарф':
        if random.randrange(1, 3) == 1:
            return 'коричневый'
        else:
            return 'верблюд'
    if thing == 'зонт':
        if random.randrange(1, 3) == 1:
            return 'сила'
        else:
            return 'синий'
    if thing == 'купальник':
        if random.randrange(1, 3) == 1:
            return 'голубой'
        else:
            return 'смотреть'
    if thing == 'плед':
        if random.randrange(1, 3) == 1:
            return 'мех'
        else:
            return 'пить'


def count_occurrences(sheet, useful_rows):
    rows = []
    occurrences = []

    for index, row in enumerate(sheet.iter_rows(max_col=5, min_row=2)):
        sublist = []
        for cell in useful_rows:
            sublist.append(row[cell].value)

        if sublist not in rows:
            rows.append(sublist)
            occurrences.append(1)
        else:
            for i in range(len(rows)):
                if rows[i] == sublist:
                    occurrences[i] += 1
                    break

    return occurrences, rows


def count_k_anonymity(occurrences):

    return min(occurrences)


def find_not_secure_rows(k, occurrences, rows):

    not_secure_rows = []
    temp = rows.copy()

    for i, occurrence in enumerate(occurrences.copy()):
        if occurrence < k:
            not_secure_rows.append(temp[i])
            occurrences.remove(occurrence)
            rows.remove(temp[i])

    return not_secure_rows, occurrences, rows


def local_suppression(sheet, single_rows, useful_rows):

    for i, row in enumerate(sheet.iter_rows(max_col=5, min_row=2)):
        sublist = []
        for cell in useful_rows:
            sublist.append(row[cell].value)

        if sublist in single_rows:
            sheet.move_range(f'A{i+3}:G{sheet.max_row + 1}', rows=-1)

    return None


def ask_about_column(useful_rows, question, column):

    print(question)
    answer = input()
    if answer == 'y':
        useful_rows.append(column)

    return useful_rows


def get_bad_occurrences(k, occurrences):

    bad_occurrences = []
    every_bad_occurrence = []
    amount_of_bad_rows = 0

    for i, occurrence in enumerate(occurrences):
        if occurrence < k and occurrence not in bad_occurrences:
            bad_occurrences.append(occurrences[i])
            every_bad_occurrence.append(occurrences[i])
            amount_of_bad_rows += occurrence
        if occurrence < k and occurrence in bad_occurrences:
            amount_of_bad_rows += occurrence
            for index, occ in enumerate(bad_occurrences):
                if occ == occurrence:
                    every_bad_occurrence[index] += occ
                    break

    return bad_occurrences, every_bad_occurrence,  amount_of_bad_rows


def get_single_rows(occurrences, rows):

    single_rows = []

    for i, occurrence in enumerate(occurrences):
        if occurrence == 1:
            single_rows.append(rows[i])

    return single_rows


def main():

    ifcount = True

    while ifcount:

        print('Обезличить датасет? y/n')
        answer = input()
        if answer == 'y' or answer == 'n':
            ifcount = False

    if answer == 'y':

        workbook = load_workbook(filename="xlsx/adv.xlsx")
        sheet = workbook.active

        codes = random.sample(range(1, 51), 50)
        platforms_for_masking = []

        for cell_number in range(2, sheet.max_row + 1):

            email = sheet[f'A{cell_number}'].value
            email = mask_email(email)
            sheet[f'A{cell_number}'] = email

            platform = sheet[f'C{cell_number}'].value
            if platform not in platforms_for_masking:
                platforms_for_masking.append(platform)
            code = alias_of_site(platform, codes, platforms_for_masking)
            sheet[f'C{cell_number}'] = code

            amount_of_ads = sheet[f'E{cell_number}'].value
            amount_of_ads = local_for_amount_of_ads(amount_of_ads)
            sheet[f'E{cell_number}'] = amount_of_ads

            adv_time = sheet[f'F{cell_number}'].value
            adv_time = local_for_adv_time(adv_time)
            sheet[f'F{cell_number}'] = adv_time

            product = sheet[f'G{cell_number}'].value
            product = local_for_product(product)
            sheet[f'G{cell_number}'] = product

        sheet.move_range(f'C1:C{sheet.max_row}', cols=-1)
        sheet.move_range(f'E1:E{sheet.max_row}', cols=-2)
        sheet.move_range(f'F1:F{sheet.max_row}', cols=-2)
        sheet.move_range(f'G1:G{sheet.max_row}', cols=-2)

        workbook.save(filename="xlsx/changed_adv.xlsx")

    ifcount = True

    while ifcount:

        print('Рассчитать k-anonymity? y/n')
        answer = input()
        if answer == 'y' or answer == 'n':
            ifcount = False

    if answer == 'y':

        useful_rows = []

        useful_rows = ask_about_column(useful_rows, 'Использовать email при пересчете? y/n', 0)
        useful_rows = ask_about_column(useful_rows, 'Использовать platform при пересчете? y/n', 1)
        useful_rows = ask_about_column(useful_rows, 'Использовать amount_of_ads при пересчете? y/n', 2)
        useful_rows = ask_about_column(useful_rows, 'Использовать duration при пересчете? y/n', 3)
        useful_rows = ask_about_column(useful_rows, 'Использовать product при пересчете? y/n', 4)

        workbook = load_workbook(filename="xlsx/changed_adv.xlsx")
        sheet = workbook.active

        occurrences, rows = count_occurrences(sheet, useful_rows)

        k = count_k_anonymity(occurrences)

        single_rows = get_single_rows(occurrences, rows)

        print(f'k-anonymity = {k}')

        if k == 1:
            print('Строки с k-anonymity = 1')
            for row in single_rows:
                print(row)

        ifcount = True

        while ifcount:

            print('Изменить k-anonymity с помощью локального подавления? y/n')
            answer = input()
            if answer == 'y' or answer == 'n':
                ifcount = False

        if answer == 'y':

            print('Введите минимальное желаемое k-anonymity: ')
            desired_k = int(input())

            bad_occurrences, every_bad_occurrences, amount_of_bad_rows = get_bad_occurrences(desired_k, occurrences)

            connected_occurrences = zip(bad_occurrences, every_bad_occurrences)
            new_connected_occurrences = sorted(connected_occurrences, key=lambda tup: tup[0])
            bad_occurrences = [connected_occurrences[0] for connected_occurrences in new_connected_occurrences]
            every_bad_occurrences = [connected_occurrences[1] for connected_occurrences in new_connected_occurrences]

            for i, occurrence in enumerate(bad_occurrences):
                print(f'Будет удалено k = {occurrence}, от набора составляет {every_bad_occurrences[i]/sheet.max_row}')
                if i == 5:
                    print('И так далее...')

            print(f'Процент неподходящих строк от основного набора = {amount_of_bad_rows / sheet.max_row}')

            unsecure_rows, occurrences, rows = find_not_secure_rows(desired_k, occurrences, rows)
            local_suppression(sheet, unsecure_rows, useful_rows)
            k = count_k_anonymity(occurrences)

            print(f'k-anonymity после подавления = {k}')

        workbook.save(filename="xlsx/changed_adv.xlsx")
    else:
        print('Заканчиваю работу')


if __name__ == '__main__':
    main()

